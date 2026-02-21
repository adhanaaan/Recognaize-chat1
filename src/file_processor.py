"""
File Processing Utility for Chat Upload
Handles extraction and parsing of various file types for inclusion in chat prompts.
"""

import os
import json
import csv
import logging
from typing import Optional, Dict, List
from pathlib import Path

logger = logging.getLogger(__name__)

# Try to import optional dependencies
try:
    import PyPDF2
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class FileProcessor:
    """Utility class for processing uploaded files."""
    
    # Supported file types and their handlers
    SUPPORTED_TYPES = {
        '.txt': 'process_text_file',
        '.csv': 'process_csv_file',
        '.json': 'process_json_file',
        '.pdf': 'process_pdf_file',
        '.xlsx': 'process_excel_file',
        '.xls': 'process_excel_file',
    }
    
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB max file size
    # To avoid exceeding the model context window, we cap how much
    # extracted content from any single file is sent to the LLM.
    # ~4k characters ≈ 1000 tokens, leaving plenty of room for
    # system prompt, knowledge-base context, and chat history.
    MAX_CONTENT_CHARS = 4000
    
    @staticmethod
    def process_uploaded_file(uploaded_file) -> Optional[Dict]:
        """Process an uploaded file object from Streamlit or FastAPI.

        Supports:
        - Streamlit's UploadedFile (has .getvalue() and .name)
        - FastAPI's UploadFile (has .file and .filename)
        """
        try:
            if uploaded_file is None:
                return None

            # Detect file wrapper type and normalize
            filename: str
            file_obj = uploaded_file

            # Streamlit UploadedFile
            if hasattr(uploaded_file, "getvalue") and hasattr(uploaded_file, "name"):
                raw_bytes = uploaded_file.getvalue()
                file_size = len(raw_bytes)
                filename = uploaded_file.name

            # FastAPI UploadFile
            elif hasattr(uploaded_file, "file") and hasattr(uploaded_file, "filename"):
                filename = uploaded_file.filename  # type: ignore[assignment]
                file_obj = uploaded_file.file      # type: ignore[assignment]
                # Compute size without consuming the stream
                current_pos = file_obj.tell()
                file_obj.seek(0, os.SEEK_END)
                file_size = file_obj.tell()
                file_obj.seek(current_pos)

            else:
                # Fallback: best-effort for generic file-like objects
                if hasattr(uploaded_file, "name"):
                    filename = uploaded_file.name  # type: ignore[assignment]
                else:
                    filename = "uploaded_file"

                try:
                    current_pos = uploaded_file.tell()
                    uploaded_file.seek(0, os.SEEK_END)
                    file_size = uploaded_file.tell()
                    uploaded_file.seek(current_pos)
                except Exception:
                    file_size = 0

            # Enforce max size
            if file_size > FileProcessor.MAX_FILE_SIZE:
                logger.warning(f"File {filename} exceeds max size ({file_size} bytes)")
                return None

            file_ext = Path(filename).suffix.lower()

            if file_ext not in FileProcessor.SUPPORTED_TYPES:
                logger.warning(f"Unsupported file type: {file_ext}")
                return None

            handler_name = FileProcessor.SUPPORTED_TYPES[file_ext]
            handler = getattr(FileProcessor, handler_name)

            # Handlers expect a file-like object with read/seek
            content = handler(file_obj)

            return {
                'filename': filename,
                'file_type': file_ext,
                'content': content,
                'size_bytes': file_size,
            }

        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            return None
    
    @staticmethod
    def process_text_file(uploaded_file) -> str:
        """Extract content from a text file."""
        try:
            content = uploaded_file.read().decode('utf-8')
            return content.strip()
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                uploaded_file.seek(0)
                content = uploaded_file.read().decode('latin-1')
                return content.strip()
            except Exception as e:
                logger.error(f"Failed to decode text file: {str(e)}")
                return ""
    
    @staticmethod
    def process_csv_file(uploaded_file) -> str:
        """Extract content from a CSV file and format as readable text."""
        try:
            uploaded_file.seek(0)
            text_content = uploaded_file.read().decode('utf-8')
            
            # Parse CSV
            lines = text_content.strip().split('\n')
            reader = csv.reader(lines)
            
            rows = list(reader)
            if not rows:
                return ""
            
            # Format as readable table
            formatted = "CSV Data:\n"
            formatted += "=" * 50 + "\n"
            
            # Add header
            if rows:
                header = rows[0]
                formatted += " | ".join(header) + "\n"
                formatted += "-" * 50 + "\n"
                
                # Add data rows (limit to first 100 rows)
                for row in rows[1:101]:
                    formatted += " | ".join(str(cell) for cell in row) + "\n"
                
                if len(rows) > 101:
                    formatted += f"\n... and {len(rows) - 101} more rows\n"
            
            return formatted.strip()
        except Exception as e:
            logger.error(f"Failed to process CSV file: {str(e)}")
            return ""
    
    @staticmethod
    def process_json_file(uploaded_file) -> str:
        """Extract content from a JSON file and format as readable text."""
        try:
            uploaded_file.seek(0)
            data = json.load(uploaded_file)
            
            # Pretty print with indentation
            formatted = json.dumps(data, indent=2)
            
            # Truncate if too long (keep first 5000 chars)
            if len(formatted) > 5000:
                formatted = formatted[:5000] + "\n\n[... content truncated ...]"
            
            return formatted
        except Exception as e:
            logger.error(f"Failed to process JSON file: {str(e)}")
            return ""
    
    @staticmethod
    def process_pdf_file(uploaded_file) -> str:
        """Extract text content from a PDF file."""
        if not PDF_AVAILABLE:
            return "[PDF support requires PyPDF2. Install with: pip install PyPDF2]"
        
        try:
            uploaded_file.seek(0)
            pdf_reader = PyPDF2.PdfReader(uploaded_file)

            text_content = ""
            # Read all pages so that downstream summarization
            # has access to the full report content.
            num_pages = len(pdf_reader.pages)

            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                page_text = page.extract_text() or ""
                text_content += f"\n--- Page {page_num + 1} ---\n"
                text_content += page_text

            text_content = text_content.strip()

            # If PyPDF2 couldn't extract meaningful text (e.g. scanned PDF),
            # fall back to OpenAI vision-based OCR.
            if len(text_content.replace("-", "").strip()) < 200:
                try:
                    # Imported as a top-level module because backend_app adds src/ to sys.path
                    from pdf_vision_extractor import extract_text_from_pdf_with_vision

                    uploaded_file.seek(0)
                    pdf_bytes = uploaded_file.read()
                    # Use a generous page limit so we effectively cover the
                    # entire ReCOGnAIze report while still having a hard cap
                    # to avoid pathological huge PDFs.
                    ocr_text = extract_text_from_pdf_with_vision(pdf_bytes, max_pages=50)
                    if ocr_text:
                        text_content = ocr_text.strip()
                except Exception as ocr_err:
                    logger.error(f"Vision OCR fallback failed: {ocr_err}")

            # More generous safeguard on length for internal
            # processing; the chat prompt will use a separate
            # summarization layer for long PDFs.
            max_pdf_chars = 20000
            if len(text_content) > max_pdf_chars:
                text_content = text_content[:max_pdf_chars] + "\n\n[... PDF content truncated for internal processing ...]"

            return text_content
        except Exception as e:
            logger.error(f"Failed to process PDF file: {str(e)}")
            return ""
    
    @staticmethod
    def process_excel_file(uploaded_file) -> str:
        """Extract content from an Excel file."""
        if not EXCEL_AVAILABLE:
            return "[Excel support requires openpyxl. Install with: pip install openpyxl]"
        
        try:
            uploaded_file.seek(0)
            workbook = openpyxl.load_workbook(uploaded_file)
            
            formatted = "Excel Data:\n"
            formatted += "=" * 50 + "\n"
            
            # Process each sheet
            for sheet_name in workbook.sheetnames[:5]:  # Limit to first 5 sheets
                sheet = workbook[sheet_name]
                formatted += f"\n--- Sheet: {sheet_name} ---\n"
                
                # Extract data (limit to first 100 rows)
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    if row_count >= 100:
                        break
                    formatted += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"
                    row_count += 1
                
                if sheet.max_row > 100:
                    formatted += f"... and {sheet.max_row - 100} more rows\n"
            
            if len(workbook.sheetnames) > 5:
                formatted += f"\n... and {len(workbook.sheetnames) - 5} more sheets"
            
            return formatted.strip()
        except Exception as e:
            logger.error(f"Failed to process Excel file: {str(e)}")
            return ""
    
    @staticmethod
    def format_file_content_for_prompt(file_data: Dict, include_metadata: bool = True) -> str:
        """
        Format processed file content for inclusion in LLM prompt.
        
        Args:
            file_data: Dictionary returned by process_uploaded_file()
            include_metadata: Whether to include filename and file type in output
            
        Returns:
            Formatted string for inclusion in prompt
        """
        if not file_data:
            return ""
        
        formatted = ""
        
        if include_metadata:
            formatted += f"File: {file_data['filename']} ({file_data['file_type']})\n"
            formatted += "=" * 60 + "\n"
        
        content = file_data['content'] or ""

        # Apply a global cap again at formatting time in case other
        # processors return very large strings.
        if len(content) > FileProcessor.MAX_CONTENT_CHARS:
            content = content[:FileProcessor.MAX_CONTENT_CHARS] + "\n\n[... file content truncated for length ...]"

        formatted += content
        
        if include_metadata:
            formatted += "\n" + "=" * 60 + "\n"
        
        return formatted
